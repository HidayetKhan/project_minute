from django.shortcuts import render
from openpyxl import load_workbook
from .models import State_Master,User_Master
from django.http import HttpResponse
import os




def import_from_excel(file):
        wb = load_workbook(file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            # print(row)
            name_state,code_state= row
            State_Master.objects.create(state_name=name_state,state_code=code_state)

        return ("done")

# import_from_excel('C:\\Users\\User\\OneDrive\\Desktop\\SQ_TASK\\ProjectMinute\\Ticket\\state.xlsx')

def user_master_file_parsel(request):
    file='C:\\Users\\User\\OneDrive\\Desktop\\SQ_TASK\\ProjectMinute\\Ticket\MOCK_DATA (1).xlsx'

    workbook = load_workbook(file)
    sheet = workbook.active
    a=0
    state = State_Master.objects.all()
    state_length = len(state)
    for row in sheet.iter_rows(min_row=2,values_only=True):
        Name,email,password,phone,city,gender,joining_date=row
        check=User_Master.objects.filter(Name=Name,email=email,phone=phone,joining_date=joining_date)
        if check:
            continue
        print(state[0].id)
        data_updating=User_Master(Name=Name,email=email,password=password,phone=phone,
        city=city, gender=gender, joining_date=joining_date, state=state[a])
        data_updating.save()
        a+=1
        if a==(state_length):
            a=0
        print(a)
    return HttpResponse("Done")
        
